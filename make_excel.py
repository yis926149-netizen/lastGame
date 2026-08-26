# -*- coding: utf-8 -*-
import sys
sys.path.insert(0, r"E:\BaiduNetdiskDownload\毕设\My project - new\.pylibs")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\Administrator\Desktop\程序需求.xlsx"

# ===== 数据 =====
data = [
    {
        "section": "人物动画",
        "cards": [
            {
                "name": "弓手",
                "type": "远程单位 · 模型待定",
                "note": "阻塞：待机 / 移动 / 攻击 / 死亡动画需优先制作 (p0)",
                "tasks": [
                    ("待机 · 移动 · 攻击 · 死亡", "必须有", "p0"),
                    ("拎起 / 扔下动画", "卡牌拖拽 → 模型切换", "p3"),
                ],
            },
            {
                "name": "剑客",
                "type": "近战单位 · 模型待定",
                "note": "阻塞：待机 / 移动 / 攻击 / 死亡动画需优先制作 (p0)",
                "tasks": [
                    ("待机 · 移动 · 攻击 · 死亡", "必须有", "p0"),
                    ("拎起 / 扔下动画", "卡牌拖拽 → 模型切换", "p3"),
                ],
            },
        ],
    },
    {
        "section": "建筑",
        "cards": [
            {
                "name": "箭塔",
                "type": "",
                "note": "",
                "tasks": [
                    ("射击", "抛物线抛射白色长方体 · 已实现", "p0"),
                    ("倒塌", "血量归零直接销毁 (暂)", "p1"),
                    ("受击", "血条 + 三档模型切换 (计划) · 目前仅血条", "p2"),
                    ("拖拽 / 生成动画", "卡牌拖拽 → 模型切换", "p3"),
                ],
            },
            {
                "name": "兵营",
                "type": "",
                "note": "",
                "tasks": [
                    ("产兵", "兵营位置凭空生成单位 · 无动画", "p1"),
                    ("倒塌", "血量归零直接销毁 (暂)", "p1"),
                    ("受击", "血条 + 三档模型切换 (计划) · 目前仅血条", "p2"),
                    ("拖拽 / 生成动画", "卡牌拖拽 → 模型切换", "p3"),
                ],
            },
            {
                "name": "金矿",
                "type": "",
                "note": "",
                "tasks": [
                    ("日常运行", "金矿车绕圈 · 目前无动画", "p1"),
                    ("倒塌", "血量归零直接销毁 (暂)", "p1"),
                    ("受击", "血条 + 三档模型切换 (计划) · 目前仅血条", "p2"),
                    ("翻地生成动画", "", "p3"),
                ],
            },
            {
                "name": "主城",
                "type": "",
                "note": "重点：主城倒塌为重要演出，需特殊处理",
                "tasks": [
                    ("倒塌", "重点演出 · 血量归零直接销毁 (暂)", "p1"),
                    ("受击", "血条 + 三档模型切换 (计划) · 目前仅血条", "p2"),
                ],
            },
        ],
    },
    {
        "section": "特效 & UI & 其他",
        "cards": [
            {
                "name": "探索地块 · 卡牌动效",
                "type": "",
                "note": "",
                "tasks": [
                    ("探索地块出现卡牌动效", "需设计", "p1"),
                ],
            },
            {
                "name": "伤害飘字",
                "type": "",
                "note": "已就绪：伤害飘字需确保暴击红色表现",
                "tasks": [
                    ("单位 / 建筑受击飘字", "暴击红色", "p0"),
                ],
            },
            {
                "name": "粒子特效",
                "type": "",
                "note": "资源：已确认可从美术资源库获取",
                "tasks": [
                    ("人物 / 建筑落地扬尘", "可复用项目组美术资源", "p3"),
                ],
            },
            {
                "name": "Unity 中文字体",
                "type": "",
                "note": "必须完成：p0 优先级，确保所有中文显示正常",
                "tasks": [
                    ("中文字体支持", "必须项", "p0"),
                ],
            },
            {
                "name": "建筑日常动画",
                "type": "",
                "note": "方案：建筑静态 + 烟雾粒子特效 + 羊群绕圈跑",
                "tasks": [
                    ("多模型动画拼接", "建筑本身不动 + 烟雾粒子 + 羊群绕圈", "p1"),
                ],
            },
        ],
    },
]

# ===== 样式 =====
PRIO_COLORS = {
    "p0": "D92D20",
    "p1": "E8590C",
    "p2": "B45309",
    "p3": "1D4ED8",
}

thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

title_font = Font(bold=True, size=16, color="0F172A")
sub_font = Font(size=11, color="64748B")
legend_font = Font(size=10, color="64748B")
header_font = Font(bold=True, size=11, color="FFFFFF")
cell_font = Font(size=11, color="1E293B")
name_font = Font(size=11, bold=True, color="0F172A")
prio_font = Font(size=11, bold=True, color="FFFFFF")

header_fill = PatternFill("solid", fgColor="1E293B")
section_fill = PatternFill("solid", fgColor="EEF2F7")
obj_fill = PatternFill("solid", fgColor="F8FAFC")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "任务清单"

# 列宽
widths = {"A": 16, "B": 20, "C": 20, "D": 28, "E": 34, "F": 9, "G": 40}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# 标题区
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "单位 & 建筑 · 开发任务汇报"
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = "更新 2026-08-21 · 版本 v0.9"
c.font = sub_font
c.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("A3:G3")
c = ws["A3"]
c.value = "优先级：p0 必须完成 · p1 重要 · p2 计划 · p3 可选"
c.font = legend_font
c.alignment = Alignment(horizontal="left", vertical="center")

# 表头
header_row = 5
headers = ["分类", "对象", "类型", "任务", "说明", "优先级", "备注"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
ws.row_dimensions[header_row].height = 22

# 数据
r = header_row + 1
for section in data:
    sec_start = r
    for card in section["cards"]:
        card_start = r
        for task_name, task_desc, prio in card["tasks"]:
            # 分类
            c = ws.cell(row=r, column=1, value=section["section"])
            c.font = cell_font
            c.fill = section_fill
            c.alignment = center
            c.border = border
            # 对象
            c = ws.cell(row=r, column=2, value=card["name"])
            c.font = name_font
            c.fill = obj_fill
            c.alignment = center
            c.border = border
            # 类型
            c = ws.cell(row=r, column=3, value=card["type"] or None)
            c.font = cell_font
            c.alignment = left
            c.border = border
            # 任务
            c = ws.cell(row=r, column=4, value=task_name)
            c.font = cell_font
            c.alignment = left
            c.border = border
            # 说明
            c = ws.cell(row=r, column=5, value=task_desc or None)
            c.font = cell_font
            c.alignment = left
            c.border = border
            # 优先级
            c = ws.cell(row=r, column=6, value=prio)
            c.font = prio_font
            c.fill = PatternFill("solid", fgColor=PRIO_COLORS[prio])
            c.alignment = center
            c.border = border
            # 备注
            c = ws.cell(row=r, column=7, value=None)
            c.font = cell_font
            c.alignment = left
            c.border = border
            r += 1

        card_end = r - 1
        # 合并 对象 / 类型 / 备注
        if card_end > card_start:
            ws.merge_cells(start_row=card_start, start_column=2, end_row=card_end, end_column=2)
            ws.merge_cells(start_row=card_start, start_column=3, end_row=card_end, end_column=3)
            ws.merge_cells(start_row=card_start, start_column=7, end_row=card_end, end_column=7)
        # 备注写入第一行
        if card["note"]:
            ws.cell(row=card_start, column=7, value=card["note"]).font = cell_font
            ws.cell(row=card_start, column=7).alignment = left

    sec_end = r - 1
    if sec_end > sec_start:
        ws.merge_cells(start_row=sec_start, start_column=1, end_row=sec_end, end_column=1)

# 冻结表头
ws.freeze_panes = "A6"

# 行高
for row in range(header_row + 1, r):
    ws.row_dimensions[row].height = 30

wb.save(OUT)
print("saved:", OUT)
